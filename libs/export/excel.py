import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles.colors import Color
from openpyxl.styles import Font, PatternFill, Border, Side
from constants.sheetnames import SheetNames

class Excel:
    """Class to export the data to excel."""
    filepath: str
    data: pd.DataFrame

    def __init__(self, filepath: str, data: pd.DataFrame) -> None:
        self.filepath = filepath
        self.data = data


    def __add_styles(self) -> bool:
        """Add the styles to the excel file."""
        try:
            workbook = load_workbook(self.filepath)
            sheet = workbook[SheetNames.ADSL]
            max_column = sheet.max_column
            max_row = sheet.max_row

            # Set the width of the columns
            sheet.column_dimensions['A'].width = 30
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 20

            # Set header styles 
            font = Font(bold=True, color=Color(rgb="FFFFFF"))
            bg = PatternFill(fill_type="solid", start_color=Color(rgb="16365C"), end_color=Color(rgb="16365C"))
            for column in range(1, max_column + 1):
                sheet.cell(row=1, column=column).font = font
                sheet.cell(row=1, column=column).fill = bg

            # Set border styles
            border = Border(left=Side(style="thin", color=Color(rgb="000000")), right=Side(style="thin", color=Color(rgb="000000")), top=Side(style="thin", color=Color(rgb="000000")), bottom=Side(style="thin", color=Color(rgb="000000")))
            for row in range(2, max_row + 1):
                for column in range(1, max_column + 1):
                    sheet.cell(row=row, column=column).border = border

            workbook.save(self.filepath)
        except Exception as error:
            print(error, __file__)
            return False
        else:
            return True
        
    def export(self) -> bool:
        """Export the data to excel.
        
        Parameters
        ----------
        data : pd.DataFrame
            Dataframe with the data to export.

        Returns
        -------
        bool
            True if the data was exported successfully, False otherwise.
        """
        try:
            self.data.to_excel(self.filepath, sheet_name=SheetNames.ADSL, index=False)
            if not self.__add_styles(): print("WARNING: Ocurred an error when adding styles to the excel file")
        except Exception as error:
            print(error, __file__)
            return False
        else:
            return True