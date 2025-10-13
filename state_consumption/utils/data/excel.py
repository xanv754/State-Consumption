import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.styles.colors import Color
from state_consumption.constants import ConsumptionStateColumns, NameColumns, cells
from state_consumption.utils.info.console import terminal
from state_consumption.utils.info.log import logger


class ExcelExport:
    """Class to export the data to excel."""
    filepath: str

    def __init__(self, filepath: str) -> None:
        self.filepath = filepath


    def _add_styles(self, sheet_name: str) -> bool:
        """Add the styles to the excel file."""
        try:
            workbook = load_workbook(self.filepath)
            sheet = workbook[sheet_name]
            max_column = sheet.max_column
            max_row = sheet.max_row

            # Set the width of the columns
            sheet.column_dimensions[cells[1]].width = 35
            for column in range(2, max_column + 1):
                max_length = len(str(sheet.cell(row=1, column=column).value))
                sheet.column_dimensions[cells[column]].width = max_length + 2

            # Set header styles 
            font = Font(bold=True, color=Color(rgb="FFFFFF"))
            bg = PatternFill(fill_type="solid", start_color=Color(rgb="16365C"), end_color=Color(rgb="16365C"))
            for column in range(1, max_column + 1):
                sheet.cell(row=1, column=column).font = font
                sheet.cell(row=1, column=column).fill = bg

            # Set border styles
            border = Border(left=Side(style="thin", color=Color(rgb="000000")), right=Side(style="thin", color=Color(rgb="000000")), top=Side(style="thin", color=Color(rgb="000000")), bottom=Side(style="thin", color=Color(rgb="000000")))
            for row in range(2, max_row + 1):
                for column in range(1, max_column + 1):
                    sheet.cell(row=row, column=column).border = border

            workbook.save(self.filepath)
        except Exception as error:
            logger.error(f"No se ha podido añadir los estilos al archivo de excel - {error}")
            terminal.print_spinner(f"[red3]ERROR: No se ha podido añadir los estilos al archivo de excel - {error}")
            return False
        else:
            return True
        
    def _add_total(self, sheet_name: str, data: pd.DataFrame) -> bool:
        """Add the total of the data to the excel file.
        
        :params sheet_name: Name of the sheet to add the data.
        :type sheet_name: str
        :params data: Dataframe with the data to add.
        :type data: DataFrame
        :returns bool: True if the data was added successfully, False otherwise.
        """
        try:
            workbook = load_workbook(self.filepath)
            sheet = workbook[sheet_name]
            max_row = sheet.max_row + 1
            border = Border(left=Side(style="thin", color=Color(rgb="000000")), right=Side(style="thin", color=Color(rgb="000000")), top=Side(style="thin", color=Color(rgb="000000")), bottom=Side(style="thin", color=Color(rgb="000000")))
            total_clients = 0
            total_consumption = 0
            total_percentage_consumption = 0

            if ConsumptionStateColumns.TOTAL_CLIENTS_ADSL in data.columns.to_list():
                total_clients_adsl = data[ConsumptionStateColumns.TOTAL_CLIENTS_ADSL].sum()
                total_clients += total_clients_adsl
                column = data.columns.get_loc(ConsumptionStateColumns.TOTAL_CLIENTS_ADSL)
                sheet.cell(row=max_row, column=column + 1).value = total_clients_adsl
                sheet.cell(row=max_row, column=column + 1).border = border
            if ConsumptionStateColumns.CONSUMPTION_ADSL in data.columns.to_list():
                total_consumption_adsl = data[ConsumptionStateColumns.CONSUMPTION_ADSL].sum()
                total_consumption += total_consumption_adsl
                column = data.columns.get_loc(ConsumptionStateColumns.CONSUMPTION_ADSL)
                sheet.cell(row=max_row, column=column + 1).value = total_consumption_adsl
                sheet.cell(row=max_row, column=column + 1).border = border
            if ConsumptionStateColumns.PERCENTAGE_CONSUMPTION_ADSL in data.columns.to_list():
                total_percentage_consumption_adsl = data[ConsumptionStateColumns.PERCENTAGE_CONSUMPTION_ADSL].sum()
                total_percentage_consumption += total_percentage_consumption_adsl
                column = data.columns.get_loc(ConsumptionStateColumns.PERCENTAGE_CONSUMPTION_ADSL)
                sheet.cell(row=max_row, column=column + 1).value = total_percentage_consumption_adsl
                sheet.cell(row=max_row, column=column + 1).border = border
            if ConsumptionStateColumns.TOTAL_CLIENTS_MDU in data.columns.to_list():
                total_clients_mdu = data[ConsumptionStateColumns.TOTAL_CLIENTS_MDU].sum()
                total_clients += total_clients_mdu
                column = data.columns.get_loc(ConsumptionStateColumns.TOTAL_CLIENTS_MDU)
                sheet.cell(row=max_row, column=column + 1).value = total_clients_mdu
                sheet.cell(row=max_row, column=column + 1).border = border
            if ConsumptionStateColumns.CONSUMPTION_MDU in data.columns.to_list():
                total_consumption_mdu = data[ConsumptionStateColumns.CONSUMPTION_MDU].sum()
                total_consumption += total_consumption_mdu
                column = data.columns.get_loc(ConsumptionStateColumns.CONSUMPTION_MDU)
                sheet.cell(row=max_row, column=column + 1).value = total_consumption_mdu
                sheet.cell(row=max_row, column=column + 1).border = border
            if ConsumptionStateColumns.PERCENTAGE_CONSUMPTION_MDU in data.columns.to_list():
                total_percentage_consumption_mdu = data[ConsumptionStateColumns.PERCENTAGE_CONSUMPTION_MDU].sum()
                total_percentage_consumption += total_percentage_consumption_mdu
                column = data.columns.get_loc(ConsumptionStateColumns.PERCENTAGE_CONSUMPTION_MDU)
                sheet.cell(row=max_row, column=column + 1).value = total_percentage_consumption_mdu
                sheet.cell(row=max_row, column=column + 1).border = border
            if ConsumptionStateColumns.TOTAL_CLIENTS_OLT in data.columns.to_list():
                total_clients_olt = data[ConsumptionStateColumns.TOTAL_CLIENTS_OLT].sum()
                total_clients += total_clients_olt
                column = data.columns.get_loc(ConsumptionStateColumns.TOTAL_CLIENTS_OLT)
                sheet.cell(row=max_row, column=column + 1).value = total_clients_olt
                sheet.cell(row=max_row, column=column + 1).border = border
            if ConsumptionStateColumns.CONSUMPTION_OLT in data.columns.to_list():
                total_consumption_olt = data[ConsumptionStateColumns.CONSUMPTION_OLT].sum()
                total_consumption += total_consumption_olt
                column = data.columns.get_loc(ConsumptionStateColumns.CONSUMPTION_OLT)
                sheet.cell(row=max_row, column=column + 1).value = total_consumption_olt
                sheet.cell(row=max_row, column=column + 1).border = border
            if ConsumptionStateColumns.PERCENTAGE_CONSUMPTION_OLT in data.columns.to_list():
                total_percentage_consumption_olt = data[ConsumptionStateColumns.PERCENTAGE_CONSUMPTION_OLT].sum()
                total_percentage_consumption += total_percentage_consumption_olt
                column = data.columns.get_loc(ConsumptionStateColumns.PERCENTAGE_CONSUMPTION_OLT)
                sheet.cell(row=max_row, column=column + 1).value = total_percentage_consumption_olt
                sheet.cell(row=max_row, column=column + 1).border = border
            if NameColumns.CONSUMPTION in data.columns.to_list():
                total_consumption_data = data[NameColumns.CONSUMPTION].sum()
                column = data.columns.get_loc(NameColumns.CONSUMPTION)
                sheet.cell(row=max_row, column=column + 1).value = total_consumption_data
                sheet.cell(row=max_row, column=column + 1).border = border
            if NameColumns.PERCENTAGE_CONSUMPTION in data.columns.to_list():
                total_percentage_consumption_data = data[NameColumns.PERCENTAGE_CONSUMPTION].sum()
                column = data.columns.get_loc(NameColumns.PERCENTAGE_CONSUMPTION)
                sheet.cell(row=max_row, column=column + 1).value = total_percentage_consumption_data
                sheet.cell(row=max_row, column=column + 1).border = border

            if total_clients != 0:
                sheet.cell(row=max_row + 4, column=1).value = "Total de Clientes"
                sheet.cell(row=max_row + 4, column=1).border = border
                sheet.cell(row=max_row + 4, column=2).value = total_clients
                sheet.cell(row=max_row + 4, column=2).border = border
            
            if total_consumption != 0:
                sheet.cell(row=max_row + 5, column=1).value = "Total de Consumo Procesado"
                sheet.cell(row=max_row + 5, column=1).border = border
                sheet.cell(row=max_row + 5, column=2).value = total_consumption
                sheet.cell(row=max_row + 5, column=2).border = border

            if total_percentage_consumption != 0:
                sheet.cell(row=max_row + 6, column=1).value = "Porcentage de Consumo Procesado"
                sheet.cell(row=max_row + 6, column=1).border = border
                sheet.cell(row=max_row + 6, column=2).value = total_percentage_consumption
                sheet.cell(row=max_row + 6, column=2).border = border

            workbook.save(self.filepath)
        except Exception as error:
            logger.error(f"No se ha podido añadir los totales al excel - {error}")
            terminal.print_spinner(f"[red3]ERROR: [default]No se ha podido añadir los totales al excel - {error}")
            return False
        else:
            return True
        
    def export(self, data: pd.DataFrame, sheet_name: str) -> bool:
        """Export the data to excel.
        
        :params data: Dataframe with the data to export.
        :type data: DataFrame
        :returns bool: True if the data was exported successfully, False otherwise.
        """
        try:
            data.to_excel(self.filepath, sheet_name=sheet_name, index=False)
            self._add_styles(sheet_name)
            self._add_total(sheet_name, data)
        except Exception as error:
            logger.error(f"No se ha podido exportar el archivo {self.filepath} - {error}")
            terminal.print_spinner(f"[red3]ERROR: [default]No se ha podido exportar el archivo {self.filepath} - {error}")
            return False
        else:
            return True
        
    def add(self, new_data: pd.DataFrame, sheet_name: str) -> bool:
        """Add the data to the excel file.
        
        :params new_data: Dataframe with the data to add.
        :type new_data: DataFrame
        :params sheet_name: Name of the sheet to add the data.
        :type sheet_name: str
        :returns bool: True if the data was added successfully, False otherwise.
        """
        try:
            with pd.ExcelWriter(self.filepath, engine='openpyxl', mode='a') as writer:
                new_data.to_excel(writer, sheet_name=sheet_name, index=False)
            self._add_styles(sheet_name)
            self._add_total(sheet_name, new_data)
        except Exception as error:
            logger.error(f"No se ha podido añadir una nueva hoja al archivo {self.filepath} - {error}")
            terminal.print_spinner(f"[red3]ERROR: [default] No se ha podido añadir una nueva hoja al archivo {self.filepath} - {error}")
            return False
        else:
            return True